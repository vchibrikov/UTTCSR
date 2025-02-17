import os
import glob
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter.font as tkFont
from openpyxl import Workbook, load_workbook

results_list = []

# Function to load .MTR file, skipping metadata rows
def load_mtr_file(file_path):
    try:
        # Load the data using a comma delimiter, and skip metadata rows
        data = pd.read_csv(file_path, delimiter = ",", skiprows = 19)  # Skip first 19 metadata rows
        
        # Clean up column names by stripping any extra spaces and handling special characters
        data.columns = [col.strip().replace('"', '') for col in data.columns]
        
        return data
    except Exception as e:
        messagebox.showerror("Error", f"Error loading file: {e}")
        return None

# Function to calculate smoothing using Savitzky-Golay filter
def apply_smoothing(elongation, force, window_length = 51, polyorder = 3):
    smoothed_force = savgol_filter(force, window_length = window_length, polyorder = polyorder)
    return smoothed_force

def write_to_excel(data):
    # Get the filename from the full path
    file_name = "uttcsr_output.xlsx"
    
    # Try to load the existing workbook, or create a new one if it doesn't exist
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "filename", "slope_1", "slope_2", "elongation_at_start_mm", "force_at_start_N",
            "elongation_at_elastic_limit_mm", "force_at_elastic_limit_N",
            "maximum_elongation_mm", "maximum_force_N", "toughness_N_mm"
        ])
        wb.save(file_name)

    # Extract only the filename from the path (no full path)
    file_name_only = os.path.basename(data[0])  # Assumes first item in data is the full path to the file
    
    # Get the active sheet (or create one if none exists)
    ws = wb.active
    # Append the data to the active sheet, replacing the full path with just the filename
    ws.append([file_name_only] + data[1:])  # Append the filename and other data
    wb.save(file_name)

class MaterialTestingApp:
    def __init__(self, root, data_list, file_names):
        self.root = root
        self.root.title("UTTCSR")

        # List of data, and an index to keep track of the current dataset
        self.data_list = data_list
        self.file_names = file_names  # List of filenames
        self.current_data_index = 0
        self.data = self.data_list[self.current_data_index]
        self.elongation = self.data["Elongation"].iloc[1:-1]  # Exclude first and last point
        self.force = self.data["Force"].iloc[1:-1]  # Exclude first and last point
        self.data_file_name = self.file_names[self.current_data_index]  # Set filename here

        # Default smoothing parameters
        self.window_length = 50
        self.polyorder = 1

        # Variables for the three points (will store elongation and force)
        self.start_point = None
        self.elastic_limit_point = None
        self.end_point = None

        # Create a frame for the plot
        self.plot_frame = tk.Frame(self.root)
        self.plot_frame.grid(row = 0, column = 0, padx = (10, 0), pady = (10, 10))

        # Create a frame for the sliders and checkbuttons
        self.control_frame = tk.Frame(self.root)
        self.control_frame.grid(row = 0, column = 1, padx = (10, 10), pady = (10, 5), sticky = "n")

        # Initially set the width of the control frame to 15% of the plot's width
        self.control_frame_width = int(self.root.winfo_width() * 0.15)
        self.control_frame.config(width = self.control_frame_width)
        
        # Configure the grid columns to allow dynamic resizing
        self.root.grid_columnconfigure(0, weight = 1)  # Allow plot area to take up remaining space
        self.root.grid_columnconfigure(1, weight = 0)  # Control frame gets fixed width

        # Plot the original data
        self.fig, self.ax = plt.subplots(figsize = (7.5, 4.5))
        self.ax.plot(self.elongation, self.force, label = "Original", color = "blue")
        self.ax.set_xlabel("Elongation (mm)")
        self.ax.set_ylabel("Force (N)")
        self.ax.legend()

        # Embed the plot into Tkinter
        self.canvas = FigureCanvasTkAgg(self.fig, master = self.plot_frame)
        self.canvas.get_tk_widget().pack()

        # Smoothing parameter controls
        self.window_label = tk.Label(self.control_frame, text = "Window length", anchor = "center")
        self.window_label.pack(pady = (0, 0), fill = "x", expand = True)

        self.window_slider = tk.Scale(self.control_frame, from_ = 1, to_ = 200, orient = "horizontal", command = self.update_plot)
        self.window_slider.set(self.window_length)
        self.window_slider.pack(pady = (0, 10), fill = "x", expand = True)

        self.poly_label = tk.Label(self.control_frame, text = "Polyorder", anchor = "center")
        self.poly_label.pack(pady = (0, 0), fill = "x", expand = True)

        self.poly_slider = tk.Scale(self.control_frame, from_  =0, to_ = 10, orient = "horizontal", command = self.update_plot)
        self.poly_slider.set(self.polyorder)
        self.poly_slider.pack(pady = (0, 10), fill = "x", expand = True)

        # Create buttons to define points
        self.start_button = tk.Button(self.control_frame, text = "Define start", command = self.define_start)
        self.start_button.pack(pady = 5, fill = "x")

        self.elastic_button = tk.Button(self.control_frame, text = "Define elastic limit", command = self.define_elastic_limit)
        self.elastic_button.pack(pady = 5, fill = "x")

        self.end_button = tk.Button(self.control_frame, text="Define end", command=self.define_end)
        self.end_button.pack(pady=5, fill="x")

        # Calculate button
        self.calculate_button = tk.Button(self.control_frame, text = "Calculate", command = self.calculate_mechanical_properties)
        self.calculate_button.pack(pady = 5, fill = "x")

        # Buttons for Next and Previous curve
        self.prev_button = tk.Button(self.control_frame, text = "Previous", command = self.previous_curve)
        self.prev_button.pack(pady = 5, fill="x")

        self.next_button = tk.Button(self.control_frame, text = "Next", command = self.next_curve)
        self.next_button.pack(pady = 5, fill = "x")

        # Checkboxes for toggling visibility of curves
        self.show_original = tk.BooleanVar(value = True)
        self.show_smoothed = tk.BooleanVar(value = True)

        # Create checkbuttons and align them to the left
        self.original_checkbox = tk.Checkbutton(self.control_frame, text = "Plot original data", variable = self.show_original, command = self.update_plot)
        self.original_checkbox.pack(pady = (0, 0), anchor = "w", padx = 0)

        self.smoothed_checkbox = tk.Checkbutton(self.control_frame, text = "Plot smoothed data", variable = self.show_smoothed, command = self.update_plot)
        self.smoothed_checkbox.pack(pady = (0, 0), anchor = "w", padx = 0)

        # Bind click event to the plot
        self.canvas.mpl_connect("button_press_event", self.on_click)

        # Initially update the plot with the smoothed data
        self.update_plot()

    def on_click(self, event):
        if event.inaxes != self.ax: 
            return

        # Get the clicked coordinates in terms of elongation (x) and force (y)
        x_click = event.xdata
        y_click = event.ydata

        # Apply smoothing to get the smoothed force values
        smoothed_force = apply_smoothing(self.elongation, self.force, self.window_length, self.polyorder)

        # Find the closest elongation value on the original data (x values)
        closest_idx = np.abs(self.elongation - x_click).argmin()

        # Get the smoothed force value at the closest elongation (y value)
        smoothed_force_at_x = smoothed_force[closest_idx]

        # Store the original x value (elongation) and the new smoothed y value (force)
        if self.start_point is None:
            self.start_point = (self.elongation.iloc[closest_idx], smoothed_force_at_x)  # Use smoothed force value
            self.ax.plot(self.start_point[0], self.start_point[1], 'go', label = "Start point")
        elif self.elastic_limit_point is None:
            self.elastic_limit_point = (self.elongation.iloc[closest_idx], smoothed_force_at_x)  # Use smoothed force value
            self.ax.plot(self.elastic_limit_point[0], self.elastic_limit_point[1], 'yo', label = "Elastic limit point")
        elif self.end_point is None:
            self.end_point = (self.elongation.iloc[closest_idx], smoothed_force_at_x)  # Use smoothed force value
            self.ax.plot(self.end_point[0], self.end_point[1], 'mo', label = "End point")

        # Redraw the plot with the marked points
        self.ax.legend()
        self.canvas.draw()

    def define_start(self):
        self.start_point = None  # Reset the start point
        messagebox.showinfo("Define start point", "Click on the plot to define the start point.")
        

    def define_elastic_limit(self):
        self.elastic_limit_point = None  # Reset the elastic limit point
        messagebox.showinfo("Define elastic limit point", "Click on the plot to define the elastic limit.")

    def define_end(self):
        self.end_point = None  # Reset the end point
        messagebox.showinfo("Define end point", "Click on the plot to define the end point.")

    def update_plot(self, event=None):
        # Get the smoothing parameters from the sliders
        self.window_length = self.window_slider.get()
        self.polyorder = self.poly_slider.get()

        if self.window_length % 2 == 0:  # Ensure window length is odd
            self.window_length += 1

        # Apply smoothing
        smoothed_force = apply_smoothing(self.elongation, self.force, self.window_length, self.polyorder)

        # Clear the previous plot
        self.ax.clear()

        # Plot the original data if the checkbox is selected
        if self.show_original.get():
            self.ax.plot(self.elongation, self.force, label = "Original", color = "blue", linestyle = "--")
        
        # Plot the smoothed data if the checkbox is selected
        if self.show_smoothed.get():
            self.ax.plot(self.elongation, smoothed_force, label = "Smoothed", color = "red")
        
        # Plot the marked points on the updated plot (use smoothed data)
        if self.start_point:
            start_x = self.start_point[0]  # Elongation (x value)
            start_y = smoothed_force[np.abs(self.elongation - start_x).argmin()]  # Find corresponding smoothed force (y value)
            self.ax.plot(start_x, start_y, 'go', label = "Start Point")
        if self.elastic_limit_point:
            elastic_x = self.elastic_limit_point[0]  # Elongation (x value)
            elastic_y = smoothed_force[np.abs(self.elongation - elastic_x).argmin()]  # Find corresponding smoothed force (y value)
            self.ax.plot(elastic_x, elastic_y, 'yo', label = "Elastic Limit")
        if self.end_point:
            end_x = self.end_point[0]  # Elongation (x value)
            end_y = smoothed_force[np.abs(self.elongation - end_x).argmin()]  # Find corresponding smoothed force (y value)
            self.ax.plot(end_x, end_y, 'mo', label = "End Point")

        # Set plot labels and title
        self.ax.set_xlabel("Elongation (mm)")
        self.ax.set_ylabel("Force (N)")
        self.ax.legend()

        # Redraw the plot
        self.canvas.draw()

    def calculate_mechanical_properties(self):
        if not self.start_point or not self.elastic_limit_point or not self.end_point:
            messagebox.showerror("Error", "Please define all points (start, elastic limit, and end) before calculating.")
            return

        # Get the elongation and force data
        elongation_range = self.elongation
        force_range = self.force

        # Randomly define 101 spans for Slope 1 (from start_point to elastic_limit_point)
        start_idx = np.abs(elongation_range - self.start_point[0]).argmin()
        elastic_idx = np.abs(elongation_range - self.elastic_limit_point[0]).argmin()
        
        slopes_1 = []
        for _ in range(101):
            rand_start = np.random.randint(start_idx, elastic_idx)
            rand_end = np.random.randint(rand_start + 1, elastic_idx + 1)  # Ensure rand_end is after rand_start
            elong_start, force_start = elongation_range.iloc[rand_start], force_range.iloc[rand_start]
            elong_end, force_end = elongation_range.iloc[rand_end], force_range.iloc[rand_end]
            slope = (force_end - force_start) / (elong_end - elong_start)
            slopes_1.append(slope)
        
        # Calculate the median of slopes for slope 1
        slope_1 = np.median(slopes_1)

        # Randomly define 101 spans for slope 2 (from elastic_limit_point to end_point)
        elastic_idx = np.abs(elongation_range - self.elastic_limit_point[0]).argmin()
        end_idx = np.abs(elongation_range - self.end_point[0]).argmin()
        
        slopes_2 = []
        for _ in range(101):
            rand_start = np.random.randint(elastic_idx, end_idx)
            rand_end = np.random.randint(rand_start + 1, end_idx + 1)  # Ensure rand_end is after rand_start
            elong_start, force_start = elongation_range.iloc[rand_start], force_range.iloc[rand_start]
            elong_end, force_end = elongation_range.iloc[rand_end], force_range.iloc[rand_end]
            slope = (force_end - force_start) / (elong_end - elong_start)
            slopes_2.append(slope)
        
        # Calculate the median of slopes for Slope 2
        slope_2 = np.median(slopes_2)

        # Calculate maximum force and elongation at maximum force
        max_force = np.max(force_range)
        max_elongation = elongation_range[np.argmax(force_range)]

        # Calculate toughness (area under the curve)
        smoothed_force = apply_smoothing(elongation_range, force_range, self.window_length, self.polyorder)
        toughness = np.trapezoid(smoothed_force, elongation_range)

        # Force and elongation at elastic limit
        force_at_elastic_limit = self.elastic_limit_point[1]
        elongation_at_elastic_limit = self.elastic_limit_point[0]

        # Prepare the results
        result = {
            "filename": self.data_file_name,
            "slope_1": slope_1,
            "slope_2": slope_2,
            "elongation_at_start_mm": self.start_point[0] - self.start_point[0],
            "force_at_start_N": self.start_point[1] - self.start_point[1],
            "elongation_at_elastic_limit_mm": elongation_at_elastic_limit - self.start_point[0],
            "force_at_elastic_limit_N": force_at_elastic_limit - self.start_point[1],
            "maximum_elongation_mm": max_elongation - self.start_point[0],
            "maximum_force_N": max_force - self.start_point[1],
            "toughness_N_mm": toughness
        }

        # Append results to a list for Excel writing
        results_list.append(result)

        # Display results in a messagebox
        messagebox.showinfo(f"slope_1: {slope_1:.6f} N/mm\n" +
                            f"slope_2: {slope_2:.6f} N/mm\n" +
                            f"elongation_at_start_mm: {self.start_point[0]:.6f} mm\n" +
                            f"force_at_start_N: {self.start_point[1]:.6f} N\n" +
                            f"maximum_force_N: {max_force:.6f} N\n" +
                            f"maximum_elongation_mm: {max_elongation:.6f} mm\n" +
                            f"toughness_N_mm: {toughness:.6f} N·mm\n" +
                            f"force_at_elastic_limit_N: {force_at_elastic_limit:.6f} N\n" +
                            f"elongation_at_elastic_limit_mm: {elongation_at_elastic_limit:.6f} mm")

        # Write to excel
        write_to_excel([result["filename"], result["slope_1"], result["slope_2"], 
                        result["elongation_at_start_mm"], result["force_at_start_N"], 
                        result["elongation_at_elastic_limit_mm"], result["force_at_elastic_limit_N"],
                        result["maximum_elongation_mm"], result["maximum_force_N"], 
                        result["toughness_N_mm"]])

    def previous_curve(self):
        if self.current_data_index > 0:
            self.current_data_index -= 1
            self.update_curve()

    def next_curve(self):
        if self.current_data_index < len(self.data_list) - 1:
            self.current_data_index += 1
            self.update_curve()

    def update_curve(self):
        self.data = self.data_list[self.current_data_index]
        self.elongation = self.data["Elongation"].iloc[1:-1]  # Exclude first and last point
        self.force = self.data["Force"].iloc[1:-1]  # Exclude first and last point
        self.data_file_name = self.file_names[self.current_data_index]  # Set filename here

        # Reset points and update plot
        self.start_point = None
        self.elastic_limit_point = None
        self.end_point = None
        self.update_plot()

# Main application code
if __name__ == "__main__":
    # Load files and prepare data
    files = filedialog.askopenfilenames(filetypes=[("MTR Files", "*.MTR")])
    data_list = []
    file_names = []
    for file in files:
        data = load_mtr_file(file)
        if data is not None:
            data_list.append(data)
            file_names.append(file)
    
    if data_list:
        # Initialize the Tkinter window
        root = tk.Tk()
        app = MaterialTestingApp(root, data_list, file_names)
        root.mainloop()
